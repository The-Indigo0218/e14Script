"""
SCRIPT 3 — Comparador de actas E-14: testigo vs Registraduría.

Lee la tabla común (SQLite), cruza por mesa, detecta discrepancias y genera:
  • Un resumen en consola.
  • Un Excel con: hoja "Resumen", hoja "Comparación" (todas las columnas lado a lado)
    y hoja "Discrepancias" (solo lo que no cuadra).

Uso:
    python comparar.py [actas.db] [salida.xlsx]
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from e14.almacen import Almacen
from e14.mesa import pares_disponibles, etiqueta_mesa
from e14.modelo import (
    FUENTE_TESTIGO,
    FUENTE_REGISTRADURIA,
    CANDIDATOS,
    CATEGORIAS_NO_CANDIDATO,
    columnas_voto,
    etiqueta_tipo_acta,
    copias_a_texto,
)
from e14.comparador import comparar, resumen, etiqueta_columna
from e14.informe import imprimir_contenido_acta
from e14.ocr import UMBRAL_REVISION

_COLS_CANDIDATOS = [f"c{n}" for n in sorted(int(k) for k in CANDIDATOS)]
_NOMBRE_C1 = CANDIDATOS[_COLS_CANDIDATOS[0][1:]][0] if len(_COLS_CANDIDATOS) == 2 else "candidato 1"
_NOMBRE_C2 = CANDIDATOS[_COLS_CANDIDATOS[1][1:]][0] if len(_COLS_CANDIDATOS) == 2 else "candidato 2"

VERDE = "C6EFCE"
ROJO = "FFC7CE"
AMARILLO = "FFEB9C"
AZUL = "1F4E78"
GRIS = "F2F2F2"
BLANCO = "FFFFFF"


def _borde():
    b = Side(style="thin", color="BFBFBF")
    return Border(left=b, right=b, top=b, bottom=b)


def _h(cell, txt, bg=AZUL, fg=BLANCO, bold=True):
    cell.value = txt
    cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(bold=bold, color=fg, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _borde()


def generar_excel(comparaciones, salida):
    wb = Workbook()

    con_par = [c for c in comparaciones if c.presente_testigo and c.presente_registraduria]
    sin_par = [c for c in comparaciones if not (c.presente_testigo and c.presente_registraduria)]
    falta_testigo = [c for c in sin_par if not c.presente_testigo]
    falta_registraduria = [c for c in sin_par if not c.presente_registraduria]

    # ── Hoja Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    r = resumen(comparaciones)
    ws.merge_cells("A1:B1")
    _h(ws["A1"], "COMPARACIÓN E-14 — TESTIGO vs REGISTRADURÍA")
    filas = [
        ("Mesas con par completo", len(con_par)),
        ("✅ Coinciden", r["COINCIDE"]),
        ("⛔ Con discrepancia", r["DISCREPANCIA"]),
        ("⏳ Sin lectura OCR", r["SIN_LECTURA"]),
        ("📋 Lectura incompleta", r["LECTURA_INCOMPLETA"]),
        ("⚠️ Falta copia testigo", len(falta_testigo)),
        ("⚠️ Falta copia registraduría", len(falta_registraduria)),
    ]
    for i, (etq, val) in enumerate(filas, start=3):
        ws.cell(i, 1, etq).font = Font(bold=True, name="Arial", size=10)
        ws.cell(i, 2, val).alignment = Alignment(horizontal="center")

    fila_sig = len(filas) + 3
    if sin_par:
        fila_sig += 1
        ws.merge_cells(f"A{fila_sig}:B{fila_sig}")
        _h(ws.cell(fila_sig, 1), "Mesas sin par completo (no entran en Comparación)", bg=AMARILLO, fg="000000")
        fila_sig += 1
        for c in sorted(sin_par, key=lambda c: c.codigo_mesa):
            falta = "Falta registraduría" if c.presente_testigo else "Falta testigo"
            ws.cell(fila_sig, 1, c.codigo_mesa).font = Font(name="Arial", size=10)
            ws.cell(fila_sig, 2, falta).alignment = Alignment(horizontal="center")
            fila_sig += 1
        fila_sig += 1

    ws.merge_cells(f"A{fila_sig}:B{fila_sig}")
    nota = ws.cell(
        fila_sig, 1,
        f"Δ Margen: positivo favorece a {_NOMBRE_C1} (c1); negativo, a {_NOMBRE_C2} (c2). "
        f"Confianza OCR < {UMBRAL_REVISION:.0%} dispara 'PENDIENTE VERIFICAR' en Verificación "
        "manual OCR; usa 'python verificar_acta.py actas.db <mesa> --fuente <testigo|registraduria> "
        "--nota \"...\"' para marcarla como verificada — queda guardado en actas.db (no en este "
        "Excel) y la próxima vez que regeneres el archivo aparecerá 'Verificado'. "
        "'Revisar' es más amplio: marca cualquier problema interno (confianza baja, suma no "
        "cuadra, alineación pobre) aunque las fuentes coincidan entre sí. % Discrepancia total "
        "mide el error global de la mesa como fracción de sus votos. '¿Cuadra? suma=total' "
        "compara la suma de candidatos+blanco+nulos+no marcados contra el \"TOTAL VOTOS DE LA "
        "MESA\" impreso en el acta; en rojo si no coinciden. La hoja Comparación tiene filtro de "
        "Excel (flechas en el encabezado): usa la de 'Confianza OCR' para ordenar de menor a "
        "mayor, la de 'Estado' para dejar solo DISCREPANCIA, y la de '¿Cuadra?' para dejar solo 'No'.",
    )
    nota.font = Font(italic=True, name="Arial", size=9, color="595959")
    nota.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    # ── Hoja Comparación (todas las columnas, testigo | reg | dif) ──
    ws2 = wb.create_sheet("Comparación")
    cols = columnas_voto()
    encabezados = [
        "Mesa", "Estado",
        "Copias en evidencia\n(Testigo)", "Votos leídos desde\n(Testigo)",
        "Copias en evidencia\n(Registr.)", "Votos leídos desde\n(Registr.)",
        "Confianza OCR\n(Testigo)", "Confianza OCR\n(Registr.)",
        f"Verificación manual OCR\n(< {UMBRAL_REVISION:.0%}) (Testigo)",
        f"Verificación manual OCR\n(< {UMBRAL_REVISION:.0%}) (Registr.)",
        "Nota verificación\n(Testigo)", "Nota verificación\n(Registr.)",
        "Revisar\n(Testigo)", "Revisar\n(Registr.)",
    ]
    for col in cols:
        et = etiqueta_columna(col)
        encabezados += [f"{et}\n(Testigo)", f"{et}\n(Registr.)", "Δ"]
    encabezados += [
        "Δ Margen\n(+c1 / -c2)",
        "Total no válidos\n(blanco+nulos+no marc.)\n(Testigo)",
        "Total no válidos\n(Registr.)", "Δ",
        "Total acta\n(Testigo)", "Total acta\n(Registr.)", "Δ",
        "Total declarado\nen el acta\n(Testigo)",
        "¿Cuadra?\nsuma=total\n(Testigo)",
        "Total declarado\nen el acta\n(Registr.)",
        "¿Cuadra?\nsuma=total\n(Registr.)",
        "% Discrepancia\ntotal",
        "KIT\n(Testigo)", "Confianza KIT\n(Testigo)",
        "KIT\n(Registr.)", "Confianza KIT\n(Registr.)",
        "CIV\n(Testigo)", "Confianza CIV\n(Testigo)",
        "CIV\n(Registr.)", "Confianza CIV\n(Registr.)",
    ]
    for j, h in enumerate(encabezados, 1):
        _h(ws2.cell(1, j), h)
    ws2.row_dimensions[1].height = 48

    def _celda_num(fila_, col_, valor, fmt=None):
        c = ws2.cell(fila_, col_, valor if valor is not None else "—")
        c.alignment = Alignment(horizontal="center")
        c.font = Font(name="Arial", size=9)
        c.border = _borde()
        if fmt and valor is not None:
            c.number_format = fmt
        return c

    fila = 2
    for comp in con_par:
        ws2.cell(fila, 1, comp.codigo_mesa).font = Font(bold=True, name="Arial", size=9)
        est = ws2.cell(fila, 2, comp.estado)
        est_bg = {
            "COINCIDE": VERDE,
            "DISCREPANCIA": ROJO,
            "FALTA_FUENTE": AMARILLO,
            "SIN_LECTURA": AMARILLO,
            "LECTURA_INCOMPLETA": AMARILLO,
        }.get(comp.estado, GRIS)
        est.fill = PatternFill("solid", start_color=est_bg)
        est.font = Font(bold=True, name="Arial", size=9)
        est.alignment = Alignment(horizontal="center")
        celdas_traz = [
            copias_a_texto(comp.copias_testigo),
            etiqueta_tipo_acta(comp.tipo_testigo),
            copias_a_texto(comp.copias_registraduria),
            etiqueta_tipo_acta(comp.tipo_registraduria),
        ]
        for off, texto in enumerate(celdas_traz):
            ce = ws2.cell(fila, 3 + off, texto)
            ce.alignment = Alignment(horizontal="center", wrap_text=True)
            ce.font = Font(name="Arial", size=9)
            ce.border = _borde()
            if off in (0, 2) and "+" in texto:
                ce.fill = PatternFill("solid", start_color=AMARILLO)

        for off, conf in enumerate((comp.confianza_testigo, comp.confianza_registraduria)):
            c = _celda_num(fila, 7 + off, conf, fmt="0%")
            if conf is not None and conf < UMBRAL_REVISION:
                c.fill = PatternFill("solid", start_color=ROJO)
                c.font = Font(bold=True, name="Arial", size=9, color="9C0006")

        confs = (comp.confianza_testigo, comp.confianza_registraduria)
        verificados = (comp.verificado_testigo, comp.verificado_registraduria)
        for off, (conf, verificado) in enumerate(zip(confs, verificados)):
            if conf is None:
                texto, color = "—", None
            elif conf >= UMBRAL_REVISION:
                texto, color = "OK", None
            elif verificado:
                texto, color = "Verificado", VERDE
            else:
                texto, color = "PENDIENTE VERIFICAR", ROJO
            c = ws2.cell(fila, 9 + off, texto)
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name="Arial", size=9, bold=(color is not None))
            c.border = _borde()
            if color:
                c.fill = PatternFill("solid", start_color=color)

        notas_verif = (comp.nota_verificacion_testigo, comp.nota_verificacion_registraduria)
        for off, nota in enumerate(notas_verif):
            c = ws2.cell(fila, 11 + off, nota or "—")
            c.alignment = Alignment(horizontal="left", wrap_text=True)
            c.font = Font(name="Arial", size=9)
            c.border = _borde()

        for off, marca in enumerate((comp.revisar_testigo, comp.revisar_registraduria)):
            c = ws2.cell(fila, 13 + off, "Sí" if marca else "No")
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name="Arial", size=9, bold=marca)
            c.border = _borde()
            if marca:
                c.fill = PatternFill("solid", start_color=AMARILLO)

        dmap = {d.columna: d for d in comp.diferencias}
        j = 15
        for col in cols:
            dc = dmap[col]
            vt, vr, d, coincide = dc.valor_testigo, dc.valor_registraduria, dc.diferencia, dc.coincide
            c_t = _celda_num(fila, j, vt)
            c_r = _celda_num(fila, j + 1, vr)
            c_d = _celda_num(fila, j + 2, d)
            if not coincide:
                for c in (c_t, c_r, c_d):
                    c.fill = PatternFill("solid", start_color=ROJO)
                c_d.font = Font(bold=True, name="Arial", size=9, color="9C0006")
            j += 3

        # ── Agregados: Δ margen, total no válidos, total acta ──
        c_margen = _celda_num(fila, j, comp.diferencia_margen)
        if comp.diferencia_margen:
            c_margen.font = Font(
                bold=True, name="Arial", size=9,
                color="9C0006" if comp.diferencia_margen < 0 else "006100",
            )
        j += 1

        nv_t, nv_r, nv_d = (
            comp.total_no_validos_testigo, comp.total_no_validos_registraduria,
            comp.diferencia_no_validos,
        )
        c_nv_t, c_nv_r, c_nv_d = _celda_num(fila, j, nv_t), _celda_num(fila, j + 1, nv_r), _celda_num(fila, j + 2, nv_d)
        if nv_d:
            for c in (c_nv_t, c_nv_r, c_nv_d):
                c.fill = PatternFill("solid", start_color=AMARILLO)
        j += 3

        tot_t, tot_r, tot_d = comp.total_testigo, comp.total_registraduria, comp.diferencia_total
        c_tot_t, c_tot_r, c_tot_d = _celda_num(fila, j, tot_t), _celda_num(fila, j + 1, tot_r), _celda_num(fila, j + 2, tot_d)
        if tot_d:
            for c in (c_tot_t, c_tot_r, c_tot_d):
                c.fill = PatternFill("solid", start_color=ROJO)
                c.font = Font(bold=True, name="Arial", size=9, color="9C0006")
        j += 3

        # ── ¿Cuadra? la suma de votos vs el "TOTAL VOTOS DE LA MESA" declarado en el acta ──
        for off, (decl, cuadra) in enumerate((
            (comp.total_declarado_testigo, comp.cuadra_testigo),
            (comp.total_declarado_registraduria, comp.cuadra_registraduria),
        )):
            c_decl = _celda_num(fila, j, decl)
            texto = "—" if cuadra is None else ("Sí" if cuadra else "No")
            c_cuadra = ws2.cell(fila, j + 1, texto)
            c_cuadra.alignment = Alignment(horizontal="center")
            c_cuadra.font = Font(name="Arial", size=9, bold=(cuadra is False))
            c_cuadra.border = _borde()
            if cuadra is False:
                c_decl.fill = PatternFill("solid", start_color=ROJO)
                c_cuadra.fill = PatternFill("solid", start_color=ROJO)
                c_cuadra.font = Font(bold=True, name="Arial", size=9, color="9C0006")
            j += 2

        c_pct = _celda_num(fila, j, comp.porcentaje_discrepancia_total, fmt="0.0%")
        if comp.porcentaje_discrepancia_total and abs(comp.porcentaje_discrepancia_total) >= 0.01:
            c_pct.fill = PatternFill("solid", start_color=ROJO)
            c_pct.font = Font(bold=True, name="Arial", size=9, color="9C0006")
        j += 1

        # ── KIT / CIV: identificadores del formulario, no son votos. Se resaltan
        # en rojo el valor si las dos fuentes lo leyeron y NO coincide (podría ser
        # evidencia de mesas distintas o un error de lectura), y la confianza si
        # quedó por debajo del umbral (el CIV en particular se imprime muy chico).
        for off, (vt, vr, ct, cr, coincide) in enumerate((
            (comp.kit_testigo, comp.kit_registraduria,
             comp.confianza_kit_testigo, comp.confianza_kit_registraduria, comp.kit_coincide),
            (comp.civ_testigo, comp.civ_registraduria,
             comp.confianza_civ_testigo, comp.confianza_civ_registraduria, comp.civ_coincide),
        )):
            base = j + off * 4
            c_t = ws2.cell(fila, base, vt or "—")
            c_r = ws2.cell(fila, base + 2, vr or "—")
            for c in (c_t, c_r):
                c.alignment = Alignment(horizontal="center")
                c.font = Font(name="Arial", size=9)
                c.border = _borde()
                if coincide is False:
                    c.fill = PatternFill("solid", start_color=ROJO)
                    c.font = Font(bold=True, name="Arial", size=9, color="9C0006")
            for off2, conf in enumerate((ct, cr)):
                c_conf = _celda_num(fila, base + 1 + off2 * 2, conf, fmt="0%")
                if conf is not None and conf < UMBRAL_REVISION:
                    c_conf.fill = PatternFill("solid", start_color=ROJO)
                    c_conf.font = Font(bold=True, name="Arial", size=9, color="9C0006")

        fila += 1
    ws2.freeze_panes = "G2"
    ws2.auto_filter.ref = ws2.dimensions
    ws2.column_dimensions["A"].width = 18
    for col in "CDEF":
        ws2.column_dimensions[col].width = 16

    # ── Hoja Trazabilidad (resumen legible por mesa) ──
    ws_t = wb.create_sheet("Trazabilidad E-14")
    for j, h in enumerate(
        ["Mesa", "Evidencia y lectura (testigo)", "Evidencia y lectura (oficial)", "Alerta"], 1
    ):
        _h(ws_t.cell(1, j), h)
    fila_t = 2
    for comp in con_par:
        alerta = ""
        if comp.alerta_copias_en_foto:
            alerta = "Varias copias en foto del testigo: revisar si los números coinciden"
        ws_t.cell(fila_t, 1, comp.codigo_mesa)
        ws_t.cell(fila_t, 2, comp.trazabilidad_testigo)
        ws_t.cell(fila_t, 3, comp.trazabilidad_registraduria)
        ws_t.cell(fila_t, 4, alerta)
        for j in range(1, 5):
            c = ws_t.cell(fila_t, j)
            c.font = Font(name="Arial", size=10)
            c.border = _borde()
            if j == 4 and alerta:
                c.fill = PatternFill("solid", start_color=AMARILLO)
        fila_t += 1
    ws_t.column_dimensions["A"].width = 18
    ws_t.column_dimensions["B"].width = 48
    ws_t.column_dimensions["C"].width = 48
    ws_t.column_dimensions["D"].width = 36

    # ── Hoja Discrepancias (solo lo que no cuadra) ──
    ws3 = wb.create_sheet("Discrepancias")
    for j, h in enumerate(["Mesa", "Columna", "Testigo", "Registraduría", "Diferencia"], 1):
        _h(ws3.cell(1, j), h)
    fila = 2
    for comp in con_par:
        if comp.estado != "DISCREPANCIA":
            continue
        for d in comp.celdas_discrepantes:
            ws3.cell(fila, 1, comp.codigo_mesa)
            ws3.cell(fila, 2, d.etiqueta)
            ws3.cell(fila, 3, d.valor_testigo if d.valor_testigo is not None else "—")
            ws3.cell(fila, 4, d.valor_registraduria if d.valor_registraduria is not None else "—")
            ws3.cell(fila, 5, d.diferencia if d.diferencia is not None else "—")
            for j in range(1, 6):
                cc = ws3.cell(fila, j)
                cc.border = _borde()
                cc.font = Font(name="Arial", size=10)
                if j >= 3:
                    cc.alignment = Alignment(horizontal="center")
            fila += 1
    for col, w in zip("ABCDE", [18, 32, 12, 14, 12]):
        ws3.column_dimensions[col].width = w

    wb.save(salida)


def main():
    db = sys.argv[1] if len(sys.argv) > 1 else "actas.db"
    salida = sys.argv[2] if len(sys.argv) > 2 else "comparacion_E14.xlsx"

    if not Path(db).exists():
        print(f"❌ No existe la base de datos: {db}")
        print("   Primero corre leer_testigos.py y leer_registraduria.py.")
        sys.exit(1)

    ambos, solo_t, solo_r = pares_disponibles("datos/testigos", "datos/registraduria")
    if ambos:
        print("Pares listos en carpetas (mismo código zona_puesto_mesa):")
        for c in sorted(ambos):
            print(f"   • {c}  ({etiqueta_mesa(c)})")
    if solo_t:
        print(f"⚠️  Solo en testigos (sin par oficial): {', '.join(sorted(solo_t))}")
    if solo_r:
        print(f"⚠️  Solo en registraduría (sin testigo): {', '.join(sorted(solo_r))}")
    if ambos or solo_t or solo_r:
        print()

    alm = Almacen(db)
    testigo = alm.leer_por_fuente(FUENTE_TESTIGO)
    registraduria = alm.leer_por_fuente(FUENTE_REGISTRADURIA)
    alm.cerrar()

    comparaciones = comparar(testigo, registraduria)
    r = resumen(comparaciones)

    print("=" * 60)
    print("  COMPARACIÓN E-14 — TESTIGO vs REGISTRADURÍA")
    print("=" * 60)
    print(f"  Mesas comparadas : {r['total']}")
    print(f"  ✅ Coinciden      : {r['COINCIDE']}")
    print(f"  ⛔ Discrepancia   : {r['DISCREPANCIA']}")
    print(f"  ⚠️  Falta fuente   : {r['FALTA_FUENTE']}")
    print(f"  ⏳ Sin lectura    : {r['SIN_LECTURA']}")
    print(f"  📋 Incompleta     : {r['LECTURA_INCOMPLETA']}")
    print("-" * 60)
    for comp in comparaciones:
        print(f"  Mesa {comp.codigo_mesa}:")
        print(f"      Jurados → {comp.trazabilidad_testigo}")
        print(f"      Oficial → {comp.trazabilidad_registraduria}")
        if comp.alerta_copias_en_foto:
            print("      ⚠️  La evidencia del testigo muestra VARIAS copias del E-14 en la misma foto.")
        if comp.estado == "DISCREPANCIA":
            print(f"      ⛔ DISCREPANCIA en votos")
            for d in comp.celdas_discrepantes:
                print(f"     - {d.etiqueta}: testigo={d.valor_testigo} "
                      f"registraduría={d.valor_registraduria} (Δ {d.diferencia})")
        elif comp.estado == "SIN_LECTURA":
            print("      ⏳ SIN LECTURA — ningún voto en testigo ni oficial (OCR falló o no corrió)")
        elif comp.estado == "LECTURA_INCOMPLETA":
            print("      📋 LECTURA INCOMPLETA — solo una fuente tiene votos o faltan columnas")
        elif comp.estado == "FALTA_FUENTE":
            falta = "registraduría" if comp.presente_testigo else "testigo"
            print(f"      ⚠️  Falta fuente '{falta}' — aún no se puede comparar votos")
            if comp.presente_registraduria and not comp.presente_testigo:
                fila_r = registraduria.get(comp.codigo_mesa)
                if fila_r:
                    imprimir_contenido_acta(
                        fila_r, titulo=f"OFICIAL en base (esperando testigo) — {comp.codigo_mesa}",
                    )
            elif comp.presente_testigo and not comp.presente_registraduria:
                fila_t = testigo.get(comp.codigo_mesa)
                if fila_t:
                    imprimir_contenido_acta(
                        fila_t, titulo=f"TESTIGO en base (esperando oficial) — {comp.codigo_mesa}",
                    )

    generar_excel(comparaciones, salida)
    print("-" * 60)
    print(f"  📊 Excel generado: {salida}")
    print("=" * 60)


if __name__ == "__main__":
    main()
